"""订单流速表 — openpyxl版"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import sqlite3
from app.constants import DB_FILE, TOTAL_ROOMS

TIME_SLOTS = [
    (1/24, "01:00"), (2/24, "02:00"), (3/24, "03:00"), (4/24, "04:00"),
    (5/24, "05:00"), (6/24, "06:00"), (7/24, "07:00"), (8/24, "08:00"),
    (0.375, "09:00"), (0.4167, "10:00"), (0.4583, "11:00"),
    (0.5, "12:00"), (0.5417, "13:00"), (0.5833, "14:00"),
    (0.625, "15:00"), (0.6667, "16:00"), (0.7083, "17:00"),
    (0.75, "18:00"), (0.7917, "19:00"), (0.8333, "20:00"),
    (0.875, "21:00"), (0.9167, "22:00"), (0.9583, "23:00"),
    (1.0, "24:00"),
]
WEEKDAYS = ['周一','周二','周三','周四','周五','周六','周日']
thin = Side(style='thin')

# 配色方案
CLR_TITLE_BG   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
CLR_TITLE_FG   = 'FFFFFF'
CLR_HEADER_BG  = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
CLR_HEADER_FG  = 'FFFFFF'
CLR_LIGHT_BG   = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
CLR_SETTLE_BG  = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')    # 淡黄(24:00行)
CLR_DAY_A      = PatternFill(start_color='F2F6FC', end_color='F2F6FC', fill_type='solid')     # 日交替A-极淡蓝
CLR_DAY_B      = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')     # 日交替B-白

def _apply_border(ws, row_start, row_end, col_start, col_end):
    """为矩形区域所有单元格加边框（含合并单元格）"""
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _style(cell, bold=False, size=10, color=None, bg=None, num_fmt=None):
    """单元格样式"""
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if bg:
        cell.fill = bg
    if num_fmt:
        cell.number_format = num_fmt

def generate_order_flow_report(start_date: str, end_date: str) -> BytesIO:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM daily_data WHERE data_date>=? AND data_date<=? ORDER BY data_date", (start_date, end_date))
    daily_rows = c.fetchall()
    c.execute("SELECT data_date,data_hour,sold_rooms,total_revenue,available_rooms,revpar,adr FROM hourly_data WHERE data_date>=? AND data_date<=? ORDER BY data_date,data_hour", (start_date, end_date))
    hourly_rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '订单流速表'

    if not daily_rows:
        ws['A1'] = '无数据'
        out = BytesIO(); wb.save(out); out.seek(0); return out

    daily = {r['data_date']: r for r in daily_rows}
    dates = sorted(daily.keys())
    # 最多导出31天
    if len(dates) > 31:
        dates = dates[:31]

    hourly_map = {}
    for r in hourly_rows:
        hourly_map.setdefault(r['data_date'], {})[r['data_hour']] = {
            'rooms': r['sold_rooms'], 'rev': r['total_revenue'],
            'avail': r['available_rooms'], 'revpar': r['revpar'], 'adr': r['adr']}

    total_cols = 1 + len(dates) * 3  # A列 + N天×3列
    last_col = get_column_letter(total_cols)
    data_start_row = 7
    data_end_row = data_start_row + len(TIME_SLOTS) - 1

    # R1: 标题
    ws.merge_cells(f'A1:{last_col}1'); c1 = ws['A1']; c1.value = '订单流速表'
    _style(c1, bold=True, size=16, color=CLR_TITLE_FG, bg=CLR_TITLE_BG)

    # R2: 房间数量
    ws.merge_cells(f'A2:{last_col}2'); c2 = ws['A2']
    c2.value = f'房间数量：{TOTAL_ROOMS}'
    _style(c2, bold=True, size=11, bg=CLR_LIGHT_BG)

    # R3-4: 起价 + 单房收益 + 平均房价（3列/天）
    for day_idx, d in enumerate(dates):
        col = 2 + day_idx * 3
        dm = daily[d]
        for j, (label, key) in enumerate([('起价','min_price'),('单房收益','revpar'),('平均房价','adr')]):
            v = dm[key]
            v_str = f'{v:.0f}' if v == int(v) else f'{v:.2f}'
            cl = ws.cell(row=3, column=col+j, value=f'{label}\n¥{v_str}')
            ws.merge_cells(start_row=3, start_column=col+j, end_row=4, end_column=col+j)
            _style(cl, bold=True, size=9, color=CLR_HEADER_FG, bg=CLR_HEADER_BG)
            cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # R5: 日期
    _style(ws.cell(row=5, column=1, value='日期'), bold=True, size=10, color=CLR_HEADER_FG, bg=CLR_HEADER_BG)
    for day_idx, d in enumerate(dates):
        col = 2 + day_idx * 3
        dt = datetime.strptime(d, '%Y-%m-%d')
        cl = ws.cell(row=5, column=col, value=f'{dt.month:02d}月{dt.day:02d}日 {WEEKDAYS[dt.weekday()]}')
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+2)
        _style(cl, bold=True, size=10, color=CLR_HEADER_FG, bg=CLR_HEADER_BG)

    # R6: 子标题
    _style(ws.cell(row=6, column=1, value='时间'), bold=True, size=9, color=CLR_HEADER_FG, bg=CLR_HEADER_BG)
    for day_idx in range(len(dates)):
        col = 2 + day_idx * 3
        for j, lbl in enumerate(['可售房间/已售房间','平均房价','单房收益']):
            _style(ws.cell(row=6, column=col+j, value=lbl), bold=True, size=9, color=CLR_HEADER_FG, bg=CLR_HEADER_BG)

    # R7-R30: 流速数据（01:00-24:00 共24个时段）
    for ri, (frac, label) in enumerate(TIME_SLOTS):
        r = data_start_row + ri
        is_settle = (label == "24:00")
        row_bg = CLR_SETTLE_BG if is_settle else None
        
        _style(ws.cell(row=r, column=1, value=label), bold=is_settle, size=10, bg=row_bg)
        for day_idx, d in enumerate(dates):
            col = 2 + day_idx * 3
            # 按日交替底色（24:00行优先淡黄）
            day_bg = row_bg if row_bg else (CLR_DAY_A if day_idx % 2 == 0 else CLR_DAY_B)
            
            hdata = hourly_map.get(d, {})
            target_h = int(frac * 24)
            best = {'rooms': 0, 'rev': 0, 'avail': TOTAL_ROOMS}
            for h in range(target_h, 0, -1):
                if h in hdata and hdata[h]['rooms'] > 0:
                    best = hdata[h]; break
            adr = round(best['rev']/best['rooms'],2) if best['rooms'] > 0 else 0
            revpar = round(best['rev']/TOTAL_ROOMS,2) if best['rev'] > 0 else 0
            for j, val in enumerate([f'{best["avail"]}/{best["rooms"]}', adr if adr > 0 else '', revpar if revpar > 0 else '']):
                fmt = '#,##0.##' if j >= 1 else None
                _style(ws.cell(row=r, column=col+j, value=val), bold=is_settle, size=10, bg=day_bg, num_fmt=fmt)

    # ---- 全区域补边框（含合并单元格）----
    _apply_border(ws, 1, 2, 1, total_cols)          # R1-R2 标题
    _apply_border(ws, 3, 4, 1, total_cols)          # R3-R4 指标
    _apply_border(ws, 5, 5, 1, total_cols)          # R5 日期
    _apply_border(ws, 6, 6, 1, total_cols)          # R6 子标题
    _apply_border(ws, data_start_row, data_end_row, 1, total_cols)  # R7+ 数据

    # 行高
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    for r in range(3, 7):
        ws.row_dimensions[r].height = 28
    for r in range(data_start_row, data_end_row + 1):
        ws.row_dimensions[r].height = 20

    # 列宽
    ws.column_dimensions['A'].width = 9
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # 冻结窗格
    ws.freeze_panes = f'A{data_start_row}'

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
