"""
数据查询与编辑接口
小时、日、月、季、年数据的CRUD
"""
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import os

from app.db.database import db_transaction
from app.schemas.common import ApiResponse
from app.core.calculator import calc_hourly, calc_daily, calc_monthly
from app.utils.validator import (
    validate_date, validate_hour, validate_hourly_data, validate_daily_data
)
from app.constants import DataSource, MAX_HOUR
from app.core.scheduler import (
    daily_aggregate,
    monthly_aggregate_for_month,
    quarterly_aggregate_for_quarter,
    yearly_aggregate_for_year,
)
from datetime import datetime as dt_mod
from app.utils.audit import audit_log
from app.api.system_api import get_current_user, require_role


def _log(action: str, detail: str = ""):
    """快捷审计日志"""
    user = get_current_user()
    audit_log(user["name"], user["role"], action, detail)


def _trigger_daily_recalc(date_str: str) -> None:
    """更新24点数据后触发日报月报重算"""
    dt = dt_mod.strptime(date_str, "%Y-%m-%d")
    daily_aggregate(date_str)
    monthly_aggregate_for_month(dt.year, dt.month)

router = APIRouter(prefix="/api", tags=["数据接口"])


class HourlyDataUpdateRequest(BaseModel):
    data_date: str
    data_hour: int
    sold_rooms: Optional[int] = None
    min_price: Optional[float] = None
    total_revenue: Optional[float] = None


class DailyDataUpdateRequest(BaseModel):
    data_date: str
    sold_rooms: Optional[int] = None
    min_price: Optional[float] = None
    total_revenue: Optional[float] = None


# ------------------------------ 小时数据 ------------------------------
@router.get("/hourly/list", summary="获取指定日期小时数据")
def get_hourly_list(date: str):
    if not validate_date(date):
        raise HTTPException(status_code=400, detail="日期格式错误，应为YYYY-MM-DD")
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        # 查询已有数据
        cursor.execute("""
            SELECT * FROM hourly_data WHERE data_date = ? ORDER BY data_hour
        """, (date,))
        exist_data = {row["data_hour"]: dict(row) for row in cursor.fetchall()}
        
        # 补全1-24点数据
        result = []
        for hour in range(1, MAX_HOUR + 1):
            if hour in exist_data:
                result.append(exist_data[hour])
            else:
                result.append({
                    "data_date": date,
                    "data_hour": hour,
                    "sold_rooms": 0,
                    "available_rooms": 113,
                    "occupancy_rate": 0,
                    "min_price": 0,
                    "revpar": 0,
                    "total_revenue": 0,
                    "data_source": DataSource.MANUAL_INPUT
                })
        
        return ApiResponse.success(result)


@router.post("/hourly/update", summary="更新小时数据")
def update_hourly_data(req: HourlyDataUpdateRequest):
    require_role("super_admin", "admin")
    if not validate_date(req.data_date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    if not validate_hour(req.data_hour):
        raise HTTPException(status_code=400, detail="小时必须在1-24之间")
    
    # 获取现有数据
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM hourly_data WHERE data_date = ? AND data_hour = ?
        """, (req.data_date, req.data_hour))
        exist = cursor.fetchone()
        
        # 合并数据
        data = dict(exist) if exist else {
            "data_date": req.data_date,
            "data_hour": req.data_hour,
            "sold_rooms": 0,
            "min_price": 0,
            "total_revenue": 0
        }
        
        if req.sold_rooms is not None:
            data["sold_rooms"] = req.sold_rooms
        if req.min_price is not None:
            data["min_price"] = req.min_price
        if req.total_revenue is not None:
            data["total_revenue"] = req.total_revenue
        
        # 校验数据
        valid, msg = validate_hourly_data(data)
        if not valid:
            raise HTTPException(status_code=400, detail=msg)
        
        # 重新计算
        calc_result = calc_hourly(data["sold_rooms"], data["total_revenue"], data["min_price"])
        data.update(calc_result)
        data["data_source"] = DataSource.MANUAL_EDIT
        
        # 写入数据库
        cursor.execute("""
            INSERT OR REPLACE INTO hourly_data
            (data_date, data_hour, sold_rooms, available_rooms, occupancy_rate, min_price, revpar, total_revenue, adr, data_source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data["data_date"],
            data["data_hour"],
            data["sold_rooms"],
            data["available_rooms"],
            data["occupancy_rate"],
            data["min_price"],
            data["revpar"],
            data["total_revenue"],
            data["adr"],
            data["data_source"]
        ))
    
    # 如果是24点数据，自动重算日报
    if req.data_hour == MAX_HOUR:
        _trigger_daily_recalc(req.data_date)
    
    return ApiResponse.success(data)


from fastapi import File, UploadFile
import csv as _csv, io as _io


@router.post("/hourly/batch-import", summary="CSV文件导入小时数据")
def batch_import_hourly(file: UploadFile = File(...)):
    """上传CSV文件批量导入小时数据，列: data_date,data_hour,sold_rooms,total_revenue,min_price"""
    require_role("super_admin", "admin")
    content = file.file.read().decode('utf-8-sig')
    reader = _csv.DictReader(_io.StringIO(content))
    
    updated = skipped = 0
    dates = set()
    for row in reader:
        date = row.get('data_date', '').strip()
        hour_str = row.get('data_hour', '').strip()
        if not date or not hour_str:
            skipped += 1; continue
        # 规范化日期格式为 YYYY-MM-DD
        try:
            date = dt_mod.strptime(date.replace('/', '-'), "%Y-%m-%d").strftime("%Y-%m-%d")
        except:
            skipped += 1; continue
        try:
            hour = int(hour_str)
            rooms_str = (row.get('sold_rooms') or '').strip()
            rev_str = (row.get('total_revenue') or '').strip()
            price_str = (row.get('min_price') or '').strip()
            rooms = int(float(rooms_str)) if rooms_str else 0
            rev = float(rev_str) if rev_str else 0.0
            price = float(price_str) if price_str else 0.0
            if not (1 <= hour <= 24): skipped += 1; continue
        except: skipped += 1; continue
        
        hd = calc_hourly(rooms, rev, price)
        with db_transaction() as conn:
            c = conn.cursor()
            c.execute("""INSERT OR REPLACE INTO hourly_data
                (data_date, data_hour, sold_rooms, available_rooms, occupancy_rate, min_price, revpar, total_revenue, adr, data_source)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (date, hour, hd["sold_rooms"], hd["available_rooms"], hd["occupancy_rate"],
                 hd["min_price"], hd["revpar"], hd["total_revenue"], hd["adr"], 3))
            updated += 1
        dates.add(date)
    
    # 触发日报/月报重算
    if updated > 0:
        for d in dates:
            try:
                daily_aggregate(d)
                dt = dt_mod.strptime(d, "%Y-%m-%d")
                monthly_aggregate_for_month(dt.year, dt.month)
            except: pass
    
    _log("批量导入小时数据", f"更新{updated}条, 跳过{skipped}条")
    return ApiResponse.success({"updated": updated, "skipped": skipped})


@router.post("/hourly/backfill-h24", summary="用日报数据回填h24小时数据")
def backfill_h24_from_daily(date: str = None):
    """指定日期用日报数据补全缺失的h24；不传date则修复所有缺失"""
    require_role("super_admin", "admin")
    import sqlite3 as _sqlite3
    from app.constants import DB_FILE
    conn = _sqlite3.connect(DB_FILE)
    conn.row_factory = _sqlite3.Row
    c = conn.cursor()
    
    if date:
        date = dt_mod.strptime(date.replace('/', '-'), "%Y-%m-%d").strftime("%Y-%m-%d")
        c.execute("SELECT * FROM daily_data WHERE data_date=?", (date,))
        dailies = c.fetchall()
    else:
        c.execute("""
            SELECT d.* FROM daily_data d
            WHERE NOT EXISTS (SELECT 1 FROM hourly_data h WHERE h.data_date=d.data_date AND h.data_hour=24)
            ORDER BY d.data_date
        """)
        dailies = c.fetchall()
    
    filled = 0
    for d in dailies:
        # 跳过无意义空日报（售出0间且收入0）
        if d["sold_rooms"] == 0 and d["total_revenue"] == 0:
            continue
        hd = calc_hourly(
            sold_rooms=d["sold_rooms"],
            total_revenue=d["total_revenue"],
            min_price=d["min_price"] or 0
        )
        c.execute("""
            INSERT OR REPLACE INTO hourly_data
            (data_date, data_hour, sold_rooms, available_rooms, occupancy_rate, min_price, revpar, total_revenue, adr, data_source)
            VALUES (?, 24, ?, ?, ?, ?, ?, ?, ?, 3)
        """, (d["data_date"], hd["sold_rooms"], hd["available_rooms"], hd["occupancy_rate"],
              hd["min_price"], hd["revpar"], hd["total_revenue"], hd["adr"]))
        filled += 1
    
    conn.commit()
    conn.close()
    return ApiResponse.success({"filled": filled, "dates": [d["data_date"] for d in dailies]})


class MinPriceUpdateRequest(BaseModel):
    data_date: str
    min_price: float


@router.put("/hourly/update-min-price", summary="更新指定日期所有小时的起售价格")
def update_min_price(req: MinPriceUpdateRequest):
    """更新指定日期所有小时数据的min_price，并重算日报"""
    require_role("super_admin", "admin")
    if not validate_date(req.data_date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    if req.min_price < 0:
        raise HTTPException(status_code=400, detail="价格不能为负数")
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        # 更新所有小时的min_price
        cursor.execute(
            "UPDATE hourly_data SET min_price = ?, update_time = datetime('now','localtime') WHERE data_date = ?",
            (req.min_price, req.data_date)
        )
        updated = cursor.rowcount
        if updated == 0:
            raise HTTPException(status_code=404, detail=f"日期 {req.data_date} 无小时数据")
    
    # 触发日报重算
    _trigger_daily_recalc(req.data_date)
    _log("更新起售价格", f"{req.data_date} → ¥{req.min_price}")
    return ApiResponse.success({"updated_rows": updated, "min_price": req.min_price})


@router.get("/hourly/export-min-price-template", summary="导出起售价格CSV模板")
def export_min_price_template():
    """导出所有日期的起售价格为CSV，供编辑后批量导入"""
    import csv, io
    from fastapi.responses import StreamingResponse

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["data_date", "min_price", "note"])
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT data_date, min_price FROM daily_data ORDER BY data_date")
        for row in cursor.fetchall():
            writer.writerow([row["data_date"], row["min_price"], ""])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=meituan_price_template.csv"}
    )


@router.post("/hourly/batch-update-min-price", summary="CSV批量更新起售价格")
def batch_update_min_price():
    """从 data/import/meituan_price.csv 批量更新所有日期的min_price"""
    import csv, os as _os
    from datetime import datetime as dtmod

    csv_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))), "data", "import", "meituan_price.csv")
    if not _os.path.exists(csv_path):
        raise HTTPException(status_code=400, detail=f"文件不存在: data/import/meituan_price.csv")
    
    updated = 0
    skipped = 0
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            date = row.get('data_date', '').strip()
            price_str = row.get('min_price', '').strip()
            if not date or not price_str:
                skipped += 1; continue
            try:
                price = float(price_str)
                if price <= 0:
                    skipped += 1; continue
            except ValueError:
                skipped += 1; continue
            
            with db_transaction() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE hourly_data SET min_price = ?, update_time = datetime('now','localtime') WHERE data_date = ?",
                    (price, date)
                )
                if cursor.rowcount > 0:
                    updated += 1
                    try:
                        daily_aggregate(date)
                        dt = dtmod.strptime(date, "%Y-%m-%d")
                        monthly_aggregate_for_month(dt.year, dt.month)
                    except: pass
                else:
                    skipped += 1
    
    return ApiResponse.success({"updated": updated, "skipped": skipped, "message": f"已更新{updated}天，跳过{skipped}天"})


@router.delete("/hourly/{hourly_id}", summary="删除指定小时数据")
def delete_hourly_data(hourly_id: int):
    require_role("super_admin", "admin")
    """删除单条小时数据，24点不可删除"""
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM hourly_data WHERE id = ?", (hourly_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        if row["data_hour"] == MAX_HOUR:
            raise HTTPException(status_code=400, detail="24点结算数据不可删除，请通过修改数据更新")
        cursor.execute("DELETE FROM hourly_data WHERE id = ?", (hourly_id,))
    return ApiResponse.success(None, message="删除成功")


@router.delete("/hourly/date/{date}", summary="删除指定日期全部小时数据")
def delete_hourly_by_date(date: str):
    require_role("super_admin", "admin")
    """删除某日所有小时数据，并清除关联日报"""
    require_role("super_admin", "admin")
    if not validate_date(date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM hourly_data WHERE data_date = ?", (date,))
        cursor.execute("DELETE FROM daily_data WHERE data_date = ?", (date,))
        # 触发月报重算
        from datetime import datetime as _dt
        dt = _dt.strptime(date, "%Y-%m-%d")
        monthly_aggregate_for_month(dt.year, dt.month)
    return ApiResponse.success(None, message=f"{date} 数据已删除")


@router.post("/daily/batch-delete", summary="批量删除日报")
def batch_delete_daily(dates: list[str]):
    require_role("super_admin", "admin")
    """批量删除指定日期的日报和小时数据，并重算月报"""
    deleted = 0
    affected_months = set()
    for date in dates:
        if not validate_date(date):
            continue
        with db_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM hourly_data WHERE data_date = ?", (date,))
            cursor.execute("DELETE FROM daily_data WHERE data_date = ?", (date,))
            deleted += 1
        try:
            dt = dt_mod.strptime(date, "%Y-%m-%d")
            affected_months.add((dt.year, dt.month))
        except: pass
    # 重算受影响月份
    for year, month in affected_months:
        try: monthly_aggregate_for_month(year, month)
        except: pass
    _log("批量删除日报", f"删除{deleted}天: {dates[:5]}...")
    return ApiResponse.success({"deleted": deleted})


@router.get("/daily/export-template", summary="导出日报Excel模板（含公式）")
def export_daily_template():
    """导出日报Excel模板，内置自动计算公式"""
    require_role("super_admin", "admin")
    import openpyxl, io as _io
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '日报数据'
    
    headers = ['日期', '已售房间', '累计房费', '起售价格', '剩余房间', '出租率%', '单房收益', '平均房价']
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT data_date, sold_rooms, total_revenue, min_price FROM daily_data ORDER BY data_date")
        for i, row in enumerate(cursor.fetchall(), 2):
            ws.cell(row=i, column=1, value=row['data_date'])
            ws.cell(row=i, column=2, value=row['sold_rooms'])
            ws.cell(row=i, column=3, value=row['total_revenue'])
            ws.cell(row=i, column=4, value=row['min_price'])
            ws.cell(row=i, column=5, value=f'=113-B{i}')
            ws.cell(row=i, column=6, value=f'=ROUND(B{i}/113*100,2)')
            ws.cell(row=i, column=7, value=f'=ROUND(C{i}/113,2)')
            ws.cell(row=i, column=8, value=f'=IF(B{i}>0,ROUND(C{i}/B{i},2),0)')
    
    for col in ['A','B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 16
    
    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_data.xlsx"})


@router.post("/daily/batch-import", summary="批量导入日报（支持CSV/XLSX）")
def batch_import_daily():
    """从 data/import/daily_data.csv 或 daily_data.xlsx 批量导入或更新日报"""
    import csv, os as _os
    base = _os.path.dirname(_os.path.dirname(_os.path.dirname(__file__)))
    import_dir = _os.path.join(base, "data", "import")
    
    # 优先尝试 XLSX，其次 CSV
    xlsx_path = _os.path.join(import_dir, "daily_data.xlsx")
    csv_path = _os.path.join(import_dir, "daily_data.csv")
    
    rows = []
    if _os.path.exists(xlsx_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"XLSX读取失败: {e}")
    elif _os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    else:
        raise HTTPException(status_code=400, detail="文件不存在: data/import/daily_data.xlsx 或 daily_data.csv")
    
    updated = skipped = 0
    for row in rows:
        date = str(row.get('data_date', row.get('日期', ''))).strip()
        if not date or not validate_date(date):
            skipped += 1; continue
        rooms = int(float(row.get('sold_rooms', row.get('售出房间', 0)) or 0))
        rev = float(row.get('total_revenue', row.get('累计房费', 0)) or 0)
        price = float(row.get('min_price', row.get('起售价格', 0)) or 0)
        
        with db_transaction() as conn:
            cursor = conn.cursor()
            dd = calc_daily(rooms, rev, price)
            cursor.execute("""INSERT OR REPLACE INTO daily_data 
                (data_date, sold_rooms, total_revenue, min_price, remain_rooms, occupancy_rate, revpar, adr, data_source)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (date, dd["sold_rooms"], dd["total_revenue"], dd["min_price"],
                 dd["remain_rooms"], dd["occupancy_rate"], dd["revpar"], dd["adr"],
                 DataSource.MANUAL_INPUT))
            updated += 1
    return ApiResponse.success({"updated": updated, "skipped": skipped})


# ------------------------------ 日报数据 ------------------------------
@router.get("/daily/list", summary="获取日报列表")
def get_daily_list(start_date: str, end_date: str):
    if not validate_date(start_date) or not validate_date(end_date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM daily_data 
            WHERE data_date >= ? AND data_date <= ?
            ORDER BY data_date DESC
        """, (start_date, end_date))
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)


@router.post("/daily/update", summary="更新日报数据")
def update_daily_data(req: DailyDataUpdateRequest):
    require_role("super_admin", "admin")
    if not validate_date(req.data_date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM daily_data WHERE data_date = ?", (req.data_date,))
        exist = cursor.fetchone()
        
        data = dict(exist) if exist else {
            "data_date": req.data_date,
            "sold_rooms": 0,
            "min_price": 0,
            "total_revenue": 0
        }
        
        if req.sold_rooms is not None:
            data["sold_rooms"] = req.sold_rooms
        if req.min_price is not None:
            data["min_price"] = req.min_price
        if req.total_revenue is not None:
            data["total_revenue"] = req.total_revenue
        
        valid, msg = validate_daily_data(data)
        if not valid:
            raise HTTPException(status_code=400, detail=msg)
        
        calc_result = calc_daily(data["sold_rooms"], data["total_revenue"], data["min_price"])
        data.update(calc_result)
        data["data_source"] = DataSource.MANUAL_EDIT
        
        cursor.execute("""
            INSERT OR REPLACE INTO daily_data
            (data_date, min_price, sold_rooms, remain_rooms, occupancy_rate, revpar, total_revenue, adr, data_source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data["data_date"],
            data["min_price"],
            data["sold_rooms"],
            data["remain_rooms"],
            data["occupancy_rate"],
            data["revpar"],
            data["total_revenue"],
            data["adr"],
            data["data_source"]
        ))
    
    # 重算月报
    _trigger_daily_recalc(req.data_date)
    return ApiResponse.success(data)


@router.delete("/daily/{date}", summary="删除日报数据")
def delete_daily_data(date: str):
    require_role("super_admin", "admin")
    if not validate_date(date):
        raise HTTPException(status_code=400, detail="日期格式错误")
    
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM daily_data WHERE data_date = ?", (date,))
        cursor.execute("DELETE FROM hourly_data WHERE data_date = ?", (date,))
    
    # 重算月报
    try:
        dt = dt_mod.strptime(date, "%Y-%m-%d")
        monthly_aggregate_for_month(dt.year, dt.month)
    except: pass
    
    return ApiResponse.success()


# ------------------------------ 月报数据 ------------------------------
@router.get("/monthly/list", summary="获取月报列表")
def get_monthly_list(year: int):
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM monthly_data WHERE data_year = ? ORDER BY data_month
        """, (year,))
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)


@router.get("/monthly/all", summary="获取全部月报数据（所有年份）")
def get_monthly_all():
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM monthly_data ORDER BY data_year, data_month")
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)


@router.get("/monthly/daily-detail", summary="获取某月每日明细")
def get_monthly_daily_detail(year: int, month: int):
    """获取指定月份的所有日报数据，用于月报图表展示"""
    month_str = f"{year}-{month:02d}"
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT data_date, total_revenue, revpar, occupancy_rate, sold_rooms
            FROM daily_data
            WHERE data_date LIKE ?
            ORDER BY data_date
        """, (f"{month_str}%",))
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)


# ------------------------------ 季报数据 ------------------------------
@router.get("/quarterly/list", summary="获取季报列表")
def get_quarterly_list(year: int):
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM quarterly_data WHERE data_year = ? ORDER BY data_quarter
        """, (year,))
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)


# ------------------------------ 年报数据 ------------------------------
@router.get("/yearly/list", summary="获取年报列表")
def get_yearly_list():
    with db_transaction() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM yearly_data ORDER BY data_year DESC")
        data = [dict(row) for row in cursor.fetchall()]
        return ApiResponse.success(data)
